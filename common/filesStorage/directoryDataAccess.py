from io import BytesIO
import os
import openpyxl

# Refactor Notes:
# 1. Use streams for large files read/write operations.
# 2. store files data in bigQuery for quick searches.
#   Then, this class implementation would be changed to BigQueryDataAccess.
# 3. This class coupled with the fact that the files are in excel formate.
#   Some other classes could extend this class and implement accordingly the logic of achieve file row.


class DirectoryDataAccess:
    def __init__(self, rootPath):
        self.rootPath = rootPath

    def storeFileInDirectory(self, dirPath: str, fileName: str, fileContent: bytes):
        fullDirPath = os.path.join(self.rootPath, dirPath)
        file = BytesIO(fileContent)
        os.makedirs(fullDirPath, exist_ok=True)
        filePath = os.path.join(fullDirPath, fileName)
        with open(filePath, "wb") as outputFile:
            outputFile.write(file.read())
        file.close()

    def applyHandlerOnDirectoryFilesRows(self, rowHandler, dirName):
        fullDirPath = os.path.join(self.rootPath, dirName)
        for subDirPath, _, files in os.walk(fullDirPath):
            for file in files:
                filePath = os.path.join(subDirPath, file)
                wb = openpyxl.load_workbook(filePath)
                for sheetName in wb.sheetnames:
                    sheet = wb[sheetName]
                    for row in sheet.iter_rows(values_only=True):
                        rowHandler(row)
                wb.close()

    def getSubDirectoryNamesContainsTermAtLeastInOneFile(self, term):
        subDirectoriesWithTerm = []
        for dirName in os.listdir(self.rootPath):
            dirPath = os.path.join(self.rootPath, dirName)
            for subDirPath, _, files in os.walk(dirPath):
                categoryContainsTerm = False
                for file in files:
                    if categoryContainsTerm:
                        break
                    filePath = os.path.join(subDirPath, file)
                    wb = openpyxl.load_workbook(filePath)
                    for sheetName in wb.sheetnames:
                        if categoryContainsTerm:
                            break
                        sheet = wb[sheetName]
                        for row in sheet.iter_rows(values_only=True):
                            if categoryContainsTerm:
                                break
                            for value in row:
                                if isinstance(value, str) and term in value:
                                    subDirectoriesWithTerm.append(
                                        os.path.basename(subDirPath))
                                    categoryContainsTerm = True
                                    break
                    wb.close()
        return subDirectoriesWithTerm
