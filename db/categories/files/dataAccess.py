from io import BytesIO
import os
from .config import FileStorageConfig
import os
import openpyxl


class FilesDataAccess:
    def store(self, dirPath: str, fileName: str, content: bytes):
        fullDirPath = os.path.join(FileStorageConfig["path"], dirPath)
        file = BytesIO(content)
        os.makedirs(fullDirPath, exist_ok=True)
        filePath = os.path.join(fullDirPath, fileName)
        # Refactor Note: use streams for large files + store in s3
        with open(filePath, "wb") as outputFile:
            outputFile.write(file.read())
        file.close()

    def applyHandlerOnDirectoryFilesRows(self, rowHandler, dirName):
        fullDirPath = os.path.join(FileStorageConfig["path"], dirName)
        for root, directories, files in os.walk(fullDirPath):
            for file in files:
                filePath = os.path.join(root, file)
                wb = openpyxl.load_workbook(filePath)
                for sheetName in wb.sheetnames:
                    sheet = wb[sheetName]
                    for row in sheet.iter_rows(values_only=True):
                        rowHandler(row)
                wb.close()
